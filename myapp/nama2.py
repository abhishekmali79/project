
from datetime import datetime
from functions import safe_int
import pandas as pd
from mysql_test import get_db_connection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import *
from PySide6.QtCore import QTimer
import sys

app = QApplication(sys.argv)

tax_rate=0.08

def safe_set(row,col,value):
    existing=table.item(row,col)

    if existing:
        existing.setText(str(value))
    else:
        table.setItem(row,col,QTableWidgetItem(str(value)))

def on_item_changed(item):
    if table.signalsBlocked():   # ensures it's not a user edit
        return
    
    print(f"User changed ({item.row()}, {item.column()}) to: {item.text()}")

    debt = safe_int(table.item(item.row(), 1).text() if table.item(item.row(),1) else 0)
    paid = safe_int(table.item(item.row(), 6).text() if table.item(item.row(),6) else 0)
    
    if item.column()==2:
        table.blockSignals(True)
        
        raw_text = item.text()

        Dict,text=dict_parser(raw_text)
        Tdebt=sum(Dict.values())
        tax=round(Tdebt*tax_rate)
        Tdebt=Tdebt+tax
        total= Tdebt + debt
        remaining=total-paid
        text=(f"{text}+{tax}")

        safe_set(item.row(), 3, Tdebt)
        safe_set(item.row(), 2, text)
        safe_set(item.row(), 4, total)
        safe_set(item.row(), 5, raw_text)
        safe_set(item.row(), 6, paid)
        safe_set(item.row(), 7, remaining)

        for col in range(table.columnCount()):
            table.resizeColumnToContents(col)

        table.blockSignals(False)

    if item.column()==1:
        table.blockSignals(True)

        debt_text = item.text().strip()
        debt = safe_int(debt_text)

        t_item=table.item(item.row(), 3)
        Tdebt = safe_int(t_item.text() if t_item else 0)

# If TOTAL is empty (at start of day), TOTAL = DEBT + Tdebt.
        old_total_item = table.item(item.row(),4)
        old_total = safe_int(old_total_item.text() if old_total_item else 0)

        if old_total == 0: 
            total = debt + Tdebt
        else:
            total = old_total + Tdebt 
            
        paid_item = table.item(item.row(), 6)
        paid = safe_int(paid_item.text() if paid_item else 0)
        remaining = total - paid

        safe_set(item.row(), 4, total)
        safe_set(item.row(), 7, remaining)

        for col in range(table.columnCount()):
            table.resizeColumnToContents(col)

        table.blockSignals(False)

    if item.column() == 6:  # PAID changed
        table.blockSignals(True)

        paid = safe_int(item.text())
        debt = safe_int(table.item(item.row(), 1).text() if table.item(item.row(), 1) else 0)
        Tdebt = safe_int(table.item(item.row(), 3).text() if table.item(item.row(), 3) else 0)
        total = safe_int(table.item(item.row(), 4).text() if table.item(item.row(), 4) else debt + Tdebt)

        remaining = total - paid
        safe_set(item.row(), 7, remaining)

        table.blockSignals(False)

def save_database():
    try:
        db=get_db_connection()
        cursor=db.cursor()
        Sql= "insert into users (NAME,DEBT,ITEMS,T_DEBT,TOTAL,PAID,REMAINING) values (%s,%s,%s,%s,%s,%s,%s)"

        for row in range(table.rowCount()):
            cursor.execute(Sql,(table.item(row,0).text(),table.item(row,1).text(),table.item(row,2).text(),table.item(row,3).text(),table.item(row,4).text(),table.item(row,6).text(),table.item(row,7).text()))
            
        db.commit()
        db.close()
        print("saved into database")

    except Exception as e:
        QMessageBox.critical(None,"Database Error", str(e))


def add_table():
    filename, _ = QFileDialog.getOpenFileName(
        None, "Select Excel File", "", "Excel Files (*.xlsx)"
    )
    if not filename:
        return
    
    df=pd.read_excel(filename,header=None)

    table.setRowCount(len(df))
    table.setColumnCount(8)
    table.setHorizontalHeaderLabels(["नाम", "उधार", "सोंदा", "आज का नामा", "टोटल","raw_items","जमा","बाकी"])

    for row in range(len(df)):
        for col in range(min(8, len(df.columns))):
            table.setItem(row, col, QTableWidgetItem(str(df.iat[row, col])))

    for row in range(table.rowCount()):
        remaining_item=table.item(row,7)
        remaining=safe_int(remaining_item.text() if remaining_item else 0)

        safe_set(row,1,remaining)
        safe_set(row,2,"")          
        safe_set(row,3,"")
        safe_set(row,4,(remaining))
        safe_set(row,5,"")
        safe_set(row,6,0)

        total=remaining+0
        remaining=total-0

        safe_set(row,4,total)
        safe_set(row,7,remaining)

    for col in range(table.columnCount()):
        table.resizeColumnToContents(col)

def load_xlsx_with_pandas(table,filename):
    df=pd.read_excel(filename,header=None)

    table.setRowCount(len(df))
    table.setColumnCount(8)
    table.setHorizontalHeaderLabels(["नाम", "उधार", "सोंदा", "आज का नामा", "टोटल","raw_items","जमा","बाकी"])

    for row in range(len(df)):
        for col in range(min(8, len(df.columns))):
            safe_set(row,col,df.iat[row, col])

def load_file():
    filename, _ = QFileDialog.getOpenFileName(
        None, "Open Excel File", "", "Excel Files (*.xlsx)"
    )
    if filename:
        load_xlsx_with_pandas(table, filename)

def find_row(name):
    for row in range (table.rowCount()):
        item = table.item(row,0)
        if item and item.text() == name:
            return row
        
    return -1

def save_to_excel():
    filename, _ = QFileDialog.getSaveFileName(
        None, "Save Excel File", "", "Excel Files (*.xlsx)"
    )
    if not filename:
        return
    
    wb=Workbook()
    ws=wb.active

    rows=table.rowCount()
    cols=table.columnCount()

    for row in range(rows):
        for col in range(cols):
            item = table.item(row,col)
            ws.cell(row=row+1,column=col+1,value=item.text() if item else "") 

    wb.save(filename)

def dict_parser(text):
    if not text:
        return{},""

    items=text.split(",")
    d={}
    formatted_items=[]

    for item in items:
        if item.strip():
            if not item:
                continue

            try:
                key,value=item.split(":")
                value = value.split("=")[0]

                x,y=value.split("*")
                result=int(x)*int(y)

                key=key.strip()

                d[key]=int(result)
                formatted_items.append(f"{key}:{x}*{y}={d[key]}")

            except:
                continue
    
    formatted_text=",".join(formatted_items)
    return d,formatted_text


def set_table():

    debt = safe_int(input_debt.text())
    paid = safe_int(input_paid.text())
    name=input_name.text()
    raw_text=input_items.text()
    Dict,text=dict_parser(raw_text)
    Tdebt=sum(Dict.values())
    tax=round(Tdebt*tax_rate)
    total= Tdebt + debt + tax
    text=(f"{text}+{tax}")
    Tdebt=Tdebt+tax
    remaining=total-paid

    if not name:
        return
    
    row=find_row(name)

    table.blockSignals(True)
    if row ==-1:
        row=table.rowCount()
        table.insertRow(row)
        table.setItem(row, 0, QTableWidgetItem(name))

    table.setItem(row, 1, QTableWidgetItem(str(debt)))
    table.setItem(row, 2, QTableWidgetItem(text))
    table.setItem(row, 3, QTableWidgetItem(str(Tdebt)))
    table.setItem(row, 4, QTableWidgetItem(str(total)))
    table.setItem(row, 5, QTableWidgetItem(raw_text))
    table.setItem(row, 6, QTableWidgetItem(str(paid)))
    table.setItem(row, 7, QTableWidgetItem(str(remaining)))

    table.blockSignals(False)

    for col in range(table.columnCount()):
        table.resizeColumnToContents(col)

    input_name.clear()
    input_debt.clear()
    input_items.clear()
    input_paid.clear()
    input_name.setFocus()

window = QWidget()
window.setWindowTitle("My First PySide6 App")

layout = QVBoxLayout()
input_layout=QHBoxLayout()

label=QLabel("hello")
layout.addWidget(label)

input_name=QLineEdit()
input_debt=QLineEdit()
input_items=QLineEdit()
input_paid=QLineEdit()

input_name.setPlaceholderText("नाम")
input_debt.setPlaceholderText("उधार")
input_items.setPlaceholderText("सोंदा")
input_paid.setPlaceholderText("जमा")

input_layout.addWidget(input_name)
input_layout.addWidget(input_debt)
input_layout.addWidget(input_items)
input_layout.addWidget(input_paid)

QTimer.singleShot(0,lambda:input_name.setFocus())
input_name.returnPressed.connect(lambda: input_debt.setFocus())
input_debt.returnPressed.connect(lambda: input_items.setFocus())
input_items.returnPressed.connect(lambda: input_paid.setFocus())
input_paid.returnPressed.connect(lambda: button.click())

table=QTableWidget()
table.setColumnCount(8)
table.setHorizontalHeaderLabels(["नाम", "उधार", "सोंदा", "आज का नामा", "टोटल","raw_items","जमा","बाकी"])
table.setColumnHidden(5,True)
table.itemChanged.connect(on_item_changed)

button = QPushButton("add to table")
save=QPushButton("save as excel")
load=QPushButton("load file")
add=QPushButton("add table")
sql=QPushButton("save into database")

sql.clicked.connect(save_database)
button.clicked.connect(set_table)
save.clicked.connect(save_to_excel)
load.clicked.connect(load_file)
add.clicked.connect(add_table)

input_layout.addWidget(save)
input_layout.addWidget(load)
input_layout.addWidget(add)
input_layout.addWidget(sql)

layout.addWidget(table)
layout.addLayout(input_layout)
layout.addWidget(button)
window.setLayout(layout)

window.show()
sys.exit(app.exec())
